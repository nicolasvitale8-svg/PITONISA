import io
import base64
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Iterable
from datetime import datetime, date
from dataclasses import dataclass
import datetime as dt
import math
import time

import numpy as np
import pandas as pd
import re
import warnings
import streamlit as st


# ---------------- Utils ----------------
# Silenciar warning específico de parseo fila a fila
warnings.filterwarnings(
    "ignore",
    message="Could not infer format, so each element will be parsed individually",
)

# Strict date helpers (D/M/YYYY)
DATE_FMT = "%d/%m/%Y"
DATE_REGEX = re.compile(r"^\s*\d{1,2}/\d{1,2}/\d{4}\s*$")

def _clean_str_series(s: pd.Series) -> pd.Series:
    """Normaliza strings: quita BOM/espacios raros y caracteres ajenos a fechas/horas.
    Conserva dígitos y separadores comunes (/, -, :, espacio, punto).
    """
    return (
        s.astype(str)
        .str.replace("\ufeff", "", regex=False)  # BOM
        .str.replace("\u200b", "", regex=False)  # zero-width space
        .str.strip()
        .str.replace(r"[^0-9/\-:\.\s]", "", regex=True)
    )

def parse_fecha_dmY(s: pd.Series) -> pd.Series:
    """Parsea exclusivamente D/M/YYYY. No usa dateutil. NaT cuando no matchea."""
    t = _clean_str_series(s)
    ok = t.str.match(DATE_REGEX)
    out = pd.Series(pd.NaT, index=t.index, dtype="datetime64[ns]")
    if ok.any():
        out.loc[ok] = pd.to_datetime(t[ok], format=DATE_FMT, errors="coerce")
    return out

def parse_hora_HM(s: pd.Series) -> pd.Series:
    t = _clean_str_series(s)
    return pd.to_datetime(t, format="%H:%M", errors="coerce")

def format_fecha_dmY(dt_series: pd.Series) -> pd.Series:
    return dt_series.dt.strftime("%d/%m/%Y")
def week_of_month(dt: pd.Timestamp) -> int:
    return int((dt.day - 1) // 7) + 1


def ensure_datetime(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return series.dt.tz_localize(None) if getattr(series.dt, "tz", None) is not None else series
    return parse_fecha_dmY(series)


def sanitize_for_display(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        s = out[col]
        try:
            if pd.api.types.is_datetime64_any_dtype(s):
                out[col] = s.dt.strftime("%d/%m/%Y")
                continue
            if s.dtype == "object":
                if s.map(lambda x: isinstance(x, (pd.Timestamp, datetime, date, np.datetime64))).any():
                    s2 = format_fecha_dmY(parse_fecha_dmY(s))
                    out[col] = s2.fillna("")
                else:
                    out[col] = s.astype(str)
                continue
            if pd.api.types.is_categorical_dtype(s):
                out[col] = s.astype(str)
                continue
            if pd.api.types.is_integer_dtype(s):
                out[col] = pd.to_numeric(s, errors="coerce").astype("Int64")
                continue
            if pd.api.types.is_float_dtype(s):
                out[col] = pd.to_numeric(s, errors="coerce")
                continue
        except Exception:
            out[col] = s.astype(str)
    return out


def _read_bytes_retry(path: Path, retries: int = 5, delay_sec: float = 0.2) -> bytes:
    last_err: Optional[Exception] = None
    for _ in range(max(1, int(retries))):
        try:
            return path.read_bytes()
        except Exception as e:
            last_err = e
            time.sleep(max(0.0, float(delay_sec)))
    # final attempt
    return path.read_bytes()


# ---------------- Audit helpers (D/M/YYYY) ----------------
def _is_dmY_series(s: pd.Series) -> pd.Series:
    t = s.astype(str).str.strip()
    return t.str.match(DATE_REGEX)


def audit_dmY(df: pd.DataFrame, sheet_name: str, col_name: str, max_examples: int = 5) -> Dict[str, object]:
    if df is None or df.empty or col_name not in df.columns:
        return {"sheet": sheet_name, "column": col_name, "total": 0, "bad": 0, "examples": []}
    col = df[col_name]
    ok = _is_dmY_series(col)
    bad_mask = (~ok).fillna(True)
    bad_count = int(bad_mask.sum())
    total = int(len(col))
    examples = col[bad_mask].dropna().astype(str).unique().tolist()[:max_examples]
    return {"sheet": sheet_name, "column": col_name, "total": total, "bad": bad_count, "examples": examples}


def audit_dmY_many(targets: List[Tuple[pd.DataFrame, str, str]]) -> Tuple[List[Dict[str, object]], pd.DataFrame]:
    report: List[Dict[str, object]] = []
    rows: List[pd.DataFrame] = []
    for df, sheet, col in targets:
        if df is None or df.empty or col not in df.columns:
            report.append({"sheet": sheet, "column": col, "total": 0, "bad": 0, "examples": []})
            continue
        res = audit_dmY(df, sheet, col)
        report.append(res)
        if res["bad"] > 0:
            mask = (~_is_dmY_series(df[col])).fillna(True)
            tmp = df.loc[mask, [col]].copy()
            tmp.insert(0, "sheet", sheet)
            tmp.insert(1, "column", col)
            tmp = tmp.rename(columns={col: "value"})
            tmp.insert(3, "row_index", tmp.index)
            rows.append(tmp)
    detail = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(columns=["sheet", "column", "value", "row_index"])
    return report, detail


def _norm_str(s: str) -> str:
    try:
        import unicodedata as _ud
        s2 = _ud.normalize("NFKC", str(s)).replace("\xa0", " ").strip()
        return " ".join(s2.split())
    except Exception:
        return str(s).strip()


def read_csv_robust(data: bytes) -> Tuple[pd.DataFrame, List[str]]:
    """Robust CSV reader: tries encodings and delimiters, cleans headers.
    Returns (df, logs)."""
    logs: List[str] = []
    import io as _io
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    delims = [",", ";", "\t", "|"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        for dlm in delims:
            try:
                df = pd.read_csv(_io.BytesIO(data), encoding=enc, sep=dlm, engine="python")
                # Clean headers
                df.columns = [_norm_str(c) for c in df.columns]
                logs.append(f"CSV leido con encoding={enc}, sep='{dlm}'")
                return df, logs
            except Exception as e:
                last_err = e
                continue
    raise ValueError(f"No se pudo leer el CSV: {last_err}")


def _to_number(val: object) -> Optional[float]:
    """Convert free-form numeric strings to float (handles 1.234,56)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if s == "":
        return None
    # Remove currency/symbols except digits, separators, minus
    import re
    s2 = re.sub(r"[^0-9,\.-]", "", s)
    # If comma appears after dot or there are both, treat comma as decimal when pattern like \d+,\d{1,}
    if re.search(r"\d,\d", s2) and s2.count(",") >= 1:
        s2 = s2.replace(".", "")
        s2 = s2.replace(",", ".")
    try:
        return float(s2)
    except Exception:
        try:
            return float(re.sub(r",", "", s2))
        except Exception:
            return None


CSV_COLMAP = {
    "fecha": ["fecha", "date", "f. fecha"],
    "item": ["ítem", "item", "producto", "nombre", "nombre ítem", "nombre item"],
    "esperado": ["esperado", "demanda", "consumo proyectado", "consumo_proyectado", "predicción", "prediccion"],
    "stock_ini": ["stock_ini", "stock inicial", "stock anterior", "stock"],
    "pedido": ["pedido"],
    "estado_ayer": ["estado_ayer", "estado ayer", "estado"],
    "retira": ["retira?", "retira", "retiro?", "retiro"],
}


def map_csv_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str], List[str]]:
    """Map CSV columns to canonical names. Returns (df_renamed, mapping, logs)."""
    logs: List[str] = []
    cols_norm = {_norm_str(c).lower(): c for c in df.columns}
    mapping: Dict[str, Optional[str]] = {k: None for k in CSV_COLMAP.keys()}
    for key, syns in CSV_COLMAP.items():
        for s in syns:
            if s in cols_norm:
                mapping[key] = cols_norm[s]
                break
    required = ["fecha", "item", "esperado"]
    missing = [k for k in required if mapping.get(k) is None]
    if missing:
        raise ValueError("Faltan columnas esenciales en CSV: " + ", ".join(missing))
    # Build rename dict
    ren = {}
    for k, src in mapping.items():
        if src is not None:
            canon = {
                "fecha": "Fecha",
                "item": "Item",
                "esperado": "Esperado",
                "stock_ini": "Stock_ini",
                "pedido": "Pedido",
                "estado_ayer": "Estado_ayer",
                # evitar '?' en nombre de columna para operaciones posteriores
                "retira": "Retira",
            }.get(k, k)
            ren[src] = canon
    out = df.rename(columns=ren)
    if mapping.get("pedido"):
        logs.append("CSV trae 'Pedido'; se ignorará y se recalculará.")
    return out, {k: (mapping[k] or "") for k in mapping}, logs


def build_orders_from_csv(
    df: pd.DataFrame,
    cfg_target: float,
    cfg_multiple: int,
    enable_am: bool,
    am_frac: float,
    target_overrides: Optional[pd.DataFrame] = None,
    enable_blocks: bool = False,
) -> Tuple[pd.DataFrame, List[str]]:
    """Compute daily orders with carryover per item/date from a normalized CSV.
    Returns (pedidos_df, logs).
    """
    logs: List[str] = []
    df = df.copy()
    # Parse dates
    df["Fecha"] = ensure_datetime(df["Fecha"]).dt.normalize()
    # Parse numbers
    df["Esperado"] = df["Esperado"].map(lambda x: _to_number(x) or 0.0)
    if "Stock_ini" in df.columns:
        df["Stock_ini"] = df["Stock_ini"].map(lambda x: _to_number(x))
    else:
        df["Stock_ini"] = None
    # Optional state normalization
    if "Estado_ayer" in df.columns:
        df["Estado_ayer"] = normalize_state_series(df["Estado_ayer"])
    # Optional Retira?
    if "Retira" in df.columns:
        df["Retira"] = df["Retira"].astype(str).str.strip().str.lower().isin(["1", "true", "si", "sí", "y", "yes"])  # type: ignore
    else:
        df["Retira"] = True  # entrega diaria por defecto

    # Dedup: aggregate by Fecha+Item
    g = df.groupby(["Fecha", "Item"], dropna=False)
    agg = g.agg({
        "Esperado": "sum",
        "Stock_ini": "first",
        "Retira": "max",
    }).reset_index()
    # Nueva implementación: soporta modo bloques usando columna 'Retira?' si enable_blocks=True.
    # También cubre modo diario (bloques de tamaño 1) si enable_blocks=False.
    # Construye pedidos con arrastre y respeta target/múltiplo/cobertura AM.
    try:
        # Prepare overrides map once (YYYY-MM-DD -> value)
        t_override_map_global: Dict[str, float] = {}
        if target_overrides is not None and not target_overrides.empty:
            dcol = target_overrides.columns[0]
            tcol = target_overrides.columns[1]
            dd = ensure_datetime(target_overrides[dcol]).dt.normalize()
            tv = pd.to_numeric(target_overrides[tcol], errors="coerce")
            for d, v in zip(dd, tv):
                if pd.notna(d) and pd.notna(v):
                    t_override_map_global[pd.Timestamp(d).date().isoformat()] = float(v)
    except Exception:
        t_override_map_global = {}
    cfg_target = max(0.0, float(cfg_target))
    cfg_multiple = max(1, int(cfg_multiple))
    am_frac = float(min(1.0, max(0.0, am_frac)))
    rows = []
    for item, sub in agg.groupby("Item", dropna=False):
        s = sub.sort_values("Fecha").reset_index(drop=True)
        fechas_seq = [pd.Timestamp(x).date() for x in s["Fecha"].tolist()]
        expected_seq = [float(x or 0.0) for x in s["Esperado"].tolist()]
        stock0 = float(_to_number(s.loc[0, "Stock_ini"]) or 0.0)
        # per-item override
        t_override_map = {k: v for k, v in t_override_map_global.items() if k in {d.isoformat() for d in fechas_seq}}
        cfg = OrderCfg(target_global=cfg_target, multiplo=cfg_multiple, enable_am=enable_am, am_frac=am_frac)
        if enable_blocks and "Retira" in s.columns:
            exc: Dict[str, int] = {}
            for d, r in zip(fechas_seq, s["Retira"].astype(bool).tolist()):
                exc[d.isoformat()] = 1 if bool(r) else 0
            cal = DeliveryCalendar(week_bits=[0, 0, 0, 0, 0, 0, 0], exceptions=exc)
            pedidos_seq, stock_fin_seq = build_orders_blocks(fechas_seq, expected_seq, stock0, t_override_map, cfg, cal)
        else:
            pedidos_seq, stock_fin_seq = build_orders_daily(fechas_seq, expected_seq, stock0, t_override_map, cfg)
        # stock_ini arrastrado
        stock_ini_seq: List[float] = [stock0] + [float(x) for x in stock_fin_seq[:-1]] if len(stock_fin_seq) > 0 else [stock0]
        for i in range(len(fechas_seq)):
            fecha = pd.Timestamp(fechas_seq[i]).normalize()
            esperado = float(expected_seq[i])
            auto = _target_auto_for_day(i, expected_seq, cfg)
            t_eff = _target_efectivo(i, fechas_seq, expected_seq, t_override_map, cfg)
            pedido_bruto = max(0.0, t_eff + esperado - stock_ini_seq[i])
            rows.append({
                "Fecha": fecha,
                "Item": item,
                "Esperado": float(esperado),
                "Stock_ini": float(stock_ini_seq[i]),
                "Target_auto": float(auto),
                "Target_efectivo": float(t_eff),
                "Pedido_bruto": float(pedido_bruto),
                "Pedido": int(max(0, int(pedidos_seq[i]))),
                "Stock_fin": float(stock_fin_seq[i]),
            })
    pedidos = pd.DataFrame(rows)
    return pedidos, logs
    # Per item, sort by date and compute orders
    cfg_target = max(0.0, float(cfg_target))
    cfg_multiple = max(1, int(cfg_multiple))
    am_frac = float(min(1.0, max(0.0, am_frac)))
    rows = []
    for item, sub in agg.groupby("Item", dropna=False):
        s = sub.sort_values("Fecha").reset_index(drop=True)
        # Build next-day expected for AM coverage
        s["Esperado_next"] = s["Esperado"].shift(-1).fillna(0.0)
        stock_cur = _to_number(s.loc[0, "Stock_ini"]) or 0.0
        for i, r in s.iterrows():
            fecha = pd.Timestamp(r["Fecha"]).normalize()
            esperado = float(r["Esperado"]) or 0.0
            # Target override
            t_override = None
            if target_overrides is not None and not target_overrides.empty:
                try:
                    dcol = target_overrides.columns[0]
                    tcol = target_overrides.columns[1]
                    m = target_overrides
                    dser = ensure_datetime(m[dcol]).dt.normalize()
                    mask = dser == fecha
                    if mask.any():
                        t_override = float(pd.to_numeric(m.loc[mask, tcol], errors="coerce").iloc[0])
                except Exception:
                    t_override = None
            auto = (am_frac * float(s.loc[i, "Esperado_next"])) if enable_am else 0.0
            t_eff = float(t_override) if (t_override is not None and not pd.isna(t_override)) else max(cfg_target, auto)
            pedido_bruto = max(0.0, t_eff + esperado - stock_cur)
            pedido = int(np.ceil(pedido_bruto / cfg_multiple) * cfg_multiple)
            stock_fin = stock_cur + pedido - esperado
            rows.append({
                "Fecha": fecha,
                "Item": item,
                "Esperado": float(esperado),
                "Stock_ini": float(stock_cur),
                "Target_auto": float(auto),
                "Target_efectivo": float(t_eff),
                "Pedido_bruto": float(pedido_bruto),
                "Pedido": int(pedido),
                "Stock_fin": float(stock_fin),
            })
            # Arrastre
            if i + 1 < len(s):
                stock_cur = float(stock_fin)  # siguiente día usa arrastre
    pedidos = pd.DataFrame(rows)
    return pedidos, logs

def read_cfg_defaults(xls_bytes: bytes) -> Tuple[float, int, bool, float, str]:
    """Read CFG from sheet CFG if present.
    Reads:
      - CFG_TARGET (B2) float >=0
      - CFG_MULTIPLO (B3) int >=1
      - CFG_ENABLE_AM_COVERAGE (B4) bool (1/0) default True
      - CFG_AM_FRAC (B5) float in [0,1] default 0.5
    Also reads optional CFG_ITEM_FILTER (B6) as text (empty = all).
    Returns (target, multiplo, enable_am, am_frac, item_filter).
    """
    try:
        xls = pd.ExcelFile(io.BytesIO(xls_bytes))
        if "CFG" not in xls.sheet_names:
            return 0.0, 1, True, 0.5, ""
        df = xls.parse("CFG", header=None)
        target = float(pd.to_numeric(df.iloc[1, 1], errors="coerce")) if df.shape[0] > 1 and df.shape[1] > 1 else 0.0
        mult = int(pd.to_numeric(df.iloc[2, 1], errors="coerce")) if df.shape[0] > 2 and df.shape[1] > 1 else 1
        enable_am_raw = df.iloc[3, 1] if df.shape[0] > 3 and df.shape[1] > 1 else 1
        am_frac = float(pd.to_numeric(df.iloc[4, 1], errors="coerce")) if df.shape[0] > 4 and df.shape[1] > 1 else 0.5
        target = max(0.0, float(target) if pd.notna(target) else 0.0)
        mult = max(1, int(mult) if pd.notna(mult) else 1)
        try:
            enable_am = bool(int(enable_am_raw)) if pd.notna(enable_am_raw) else True
        except Exception:
            enable_am = True
        am_frac = float(am_frac) if pd.notna(am_frac) else 0.5
        am_frac = min(1.0, max(0.0, am_frac))
        item_filter = ""
        try:
            if df.shape[0] > 5 and df.shape[1] > 1:
                val = df.iloc[5, 1]
                item_filter = str(val) if pd.notna(val) else ""
        except Exception:
            item_filter = ""
        return target, mult, enable_am, am_frac, item_filter
    except Exception:
        return 0.0, 1, True, 0.5, ""


def write_cfg_to_file(
    xlsx_path: Path,
    target: float,
    multiple: int,
    enable_am: bool = True,
    am_frac: float = 0.5,
    item_filter: str = "",
    delivery_week: Optional[List[int]] = None,
) -> bool:
    """Persist Target/Múltiplo/AM Coverage into CFG and define named ranges.
    Returns True if write succeeded.
    """
    try:
        import openpyxl as ox
        from openpyxl.workbook.defined_name import DefinedName
        wb = ox.load_workbook(xlsx_path)
        sh = wb["CFG"] if "CFG" in wb.sheetnames else wb.create_sheet("CFG")
        # Ensure at least B3 exists
        sh.cell(row=2, column=2, value=max(0.0, float(target)))  # B2 CFG_TARGET
        sh.cell(row=3, column=2, value=max(1, int(multiple)))    # B3 CFG_MULTIPLO
        sh.cell(row=4, column=2, value=1 if bool(enable_am) else 0)  # B4 CFG_ENABLE_AM_COVERAGE
        sh.cell(row=5, column=2, value=min(1.0, max(0.0, float(am_frac))))  # B5 CFG_AM_FRAC
        # Optional item filter
        sh.cell(row=6, column=2, value=str(item_filter or ""))  # B6 CFG_ITEM_FILTER (text)
        # Weekly delivery calendar headers B7:H7 and default values B8:H8 (1s)
        headers = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        for j, name in enumerate(headers, start=2):
            try:
                sh.cell(row=7, column=j, value=name)
            except Exception:
                pass
        try:
            if delivery_week is not None and len(delivery_week) == 7:
                vals = [1 if int(v) == 1 else 0 for v in delivery_week]
                for j, v in enumerate(vals, start=2):
                    sh.cell(row=8, column=j, value=int(v))
            else:
                # If no values yet in B8:H8, default to ones
                existing = [sh.cell(row=8, column=j).value for j in range(2, 9)]
                if all(v is None for v in existing):
                    for j in range(2, 9):
                        sh.cell(row=8, column=j, value=1)
        except Exception:
            pass
        # Named ranges
        try:
            # Remove existing definitions if present
            names = {dn.name: dn for dn in list(wb.defined_names.definedName)}
            if "CFG_TARGET" in names:
                wb.defined_names.delete("CFG_TARGET")
            if "CFG_MULTIPLO" in names:
                wb.defined_names.delete("CFG_MULTIPLO")
            if "CFG_ENABLE_AM_COVERAGE" in names:
                wb.defined_names.delete("CFG_ENABLE_AM_COVERAGE")
            if "CFG_AM_FRAC" in names:
                wb.defined_names.delete("CFG_AM_FRAC")
            if "CFG_ITEM_FILTER" in names:
                wb.defined_names.delete("CFG_ITEM_FILTER")
            if "CFG_DELIVERY_WEEK" in names:
                wb.defined_names.delete("CFG_DELIVERY_WEEK")
            wb.defined_names.append(DefinedName(name="CFG_TARGET", attr_text="CFG!$B$2"))
            wb.defined_names.append(DefinedName(name="CFG_MULTIPLO", attr_text="CFG!$B$3"))
            wb.defined_names.append(DefinedName(name="CFG_ENABLE_AM_COVERAGE", attr_text="CFG!$B$4"))
            wb.defined_names.append(DefinedName(name="CFG_AM_FRAC", attr_text="CFG!$B$5"))
            wb.defined_names.append(DefinedName(name="CFG_ITEM_FILTER", attr_text="CFG!$B$6"))
            wb.defined_names.append(DefinedName(name="CFG_DELIVERY_WEEK", attr_text="CFG!$B$8:$H$8"))
        except Exception:
            pass
        wb.save(xlsx_path)
        return True
    except Exception:
        return False

def normalize_item_name(val: str) -> str:
    try:
        import unicodedata as _ud
        s = _ud.normalize("NFKC", str(val)).replace("\xa0", " ").strip()
        s = " ".join(s.split())
        # remove diacritics to make matching accent-insensitive
        s_no_diac = ''.join(ch for ch in _ud.normalize('NFD', s) if _ud.category(ch) != 'Mn')
        return s_no_diac.upper()
    except Exception:
        try:
            return str(val).strip().upper()
        except Exception:
            return ""


def std_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize column names to expected canonical Spanish headers.

    - Normaliza Unicode (NFKC), espacios y mayúsculas/minúsculas.
    - Renombra sin alterar datos: Fecha, Estado, Nombre, UN VENDIDA, ITEM, RUBRO,
      Stock, Saldo, Disponible, SEM / MES, D (Día con posibles problemas de encoding).
    """
    if df is None or df.empty:
        return df
    import unicodedata as _ud
    def _norm_key(s: str) -> str:
        s2 = _ud.normalize("NFKC", str(s)).replace("\xa0", " ").replace("_", " ").strip()
        s2 = " ".join(s2.split()).lower()
        return s2
    # Map normalized keys to canonical names used elsewhere in the app
    canonical = {
        "fecha": "Fecha",
        "estado": "Estado",
        "nombre": "Nombre",
        "un vendida": "UN VENDIDA",
        "item": "ITEM",
        "rubro": "RUBRO",
        "stock": "Stock",
        "saldo": "Saldo",
        "disponible": "Disponible",
        "sem / mes": "SEM / MES",
        "dia": "D",  # coincide con constante ES_VENTAS["dow"] en este archivo
    }
    ren = {}
    for c in df.columns:
        key = _norm_key(c)
        if key in canonical:
            ren[c] = canonical[key]
    out = df.rename(columns=ren) if ren else df
    # Heurística: si falta 'Estado', intentar detectar por contenido (CRIT/BAJO/OK/BUENO/TOP)
    try:
        if "Estado" not in out.columns:
            cand_cols = [c for c in out.columns if out[c].dtype == "object" or str(out[c].dtype).startswith("category")]
            best_col = None
            best_score = 0.0
            for c in cand_cols:
                s = out[c].astype(str).str.upper()
                hits = s.str.contains("CRIT|BAJO|\bOK\b|BUENO|TOP", na=False, regex=True)
                score = hits.mean() if len(s) else 0.0
                if score > best_score:
                    best_score = score
                    best_col = c
            if best_col is not None and best_score >= 0.2:  # al menos 20% de filas parecen estado
                out.rename(columns={best_col: "Estado"}, inplace=True)
    except Exception:
        pass
    # Heurística: si falta 'Fecha', detectar columna con mayoría de fechas válidas
    try:
        if "Fecha" not in out.columns:
            best_col = None
            best_score = 0.0
            for c in out.columns:
                dt = parse_fecha_dmY(out[c])
                score = dt.notna().mean() if len(out[c]) else 0.0
                if score > best_score:
                    best_score = score
                    best_col = c
            if best_col is not None and best_score >= 0.5:  # mayoría de celdas parsean como fecha
                out.rename(columns={best_col: "Fecha"}, inplace=True)
    except Exception:
        pass
    return out


# ---------------- Strict schema ----------------
ES_DIARIO = {"date": "Fecha", "item": "Nombre", "state": "Estado", "units": "UN VENDIDA"}
ES_VENTAS = {"date": "Fecha", "state": "Estado", "dow": "Día", "wom": "SEM / MES"}
STATES = ["▼▼ CRIT", "▼ BAJO", "— OK", "▲ BUENO", "▲▲ TOP"]


def _normalize_state_name(raw: str) -> str:
    """Map visual/symbol variants to canonical internal STATES.

    Accepts: "▼▼ CRIT", "▼ BAJO", "▬ OK", "▲ BUENO", "▲▲ TOP",
    and plain: "CRIT", "BAJO", "OK", "BUENO", "TOP".
    Returns one of STATES when matched; otherwise returns the trimmed input.
    """
    if raw is None:
        return ""
    s = str(raw)
    try:
        import unicodedata as _ud
        s = _ud.normalize("NFKC", s).replace("\xa0", " ").strip()
    except Exception:
        s = s.strip()
    # Drop leading decorative symbols
    while s and (s[0] in {"▲", "▼", "▬", "—", "-"} or s[:2] in {"- ", "– ", "— ", "• ", "* "}):
        s = s[1:].lstrip()
    key = s.upper()
    if "CRIT" in key:
        return STATES[0]
    if "BAJO" in key:
        return STATES[1]
    if "OK" in key:
        return STATES[2]
    if "BUENO" in key:
        return STATES[3]
    if "TOP" in key:
        return STATES[4]
    return s


def normalize_state_series(series: pd.Series) -> pd.Series:
    return series.astype(str).map(_normalize_state_name)


# ---------------- Delivery calendar and block logic ----------------
@dataclass
class DeliveryCalendar:
    # 7 ints (1/0) for [Mon..Sun]
    week_bits: List[int]
    # Exceptions by ISO date string YYYY-MM-DD -> 0/1
    exceptions: Dict[str, int]

    def __post_init__(self):
        if self.week_bits is None or len(self.week_bits) != 7:
            self.week_bits = [1, 1, 1, 1, 1, 1, 1]
        if self.exceptions is None:
            self.exceptions = {}

    def is_delivery(self, d: dt.date) -> int:
        key = d.isoformat()
        if key in self.exceptions:
            return 1 if int(self.exceptions[key]) == 1 else 0
        # weekday(): Monday=0..Sunday=6
        return 1 if int(self.week_bits[int(d.weekday())]) == 1 else 0


def _ceil_to_multiple(x: float, m: int) -> int:
    m = max(1, int(m or 1))
    if x <= 0:
        return 0
    return int(math.ceil(float(x) / m) * m)


@dataclass
class OrderCfg:
    target_global: float = 0.0
    multiplo: int = 1
    enable_am: bool = True
    am_frac: float = 0.5
    enable_blocks: Optional[bool] = None  # None => auto (según calendario/excepciones)


def _target_auto_for_day(idx: int, esperado: List[float], cfg: OrderCfg) -> float:
    if not cfg.enable_am:
        return 0.0
    nxt = esperado[idx + 1] if (idx + 1) < len(esperado) and esperado[idx + 1] is not None else 0.0
    return float(cfg.am_frac) * float(nxt)


def _target_efectivo(idx: int,
                     fechas: List[dt.date],
                     esperado: List[float],
                     target_override: Dict[str, float],
                     cfg: OrderCfg) -> float:
    key = fechas[idx].isoformat()
    if key in target_override and target_override[key] is not None:
        try:
            return float(target_override[key])
        except Exception:
            pass
    return max(float(cfg.target_global), _target_auto_for_day(idx, esperado, cfg))


def _compute_blocks(fechas: List[dt.date], cal: DeliveryCalendar) -> List[Tuple[int, int]]:
    """Return inclusive-exclusive block ranges [s, e) based on delivery markers.
    A block starts at a day with delivery=1 and ends the day before the next delivery.
    Consecutive delivery days produce 1-day blocks.
    """
    n = len(fechas)
    if n == 0:
        return []
    starts = [i for i in range(n) if cal.is_delivery(fechas[i]) == 1]
    if 0 not in starts:
        starts = sorted(set([0] + starts))
    blocks: List[Tuple[int, int]] = []
    for si, s in enumerate(starts):
        e = starts[si + 1] if (si + 1) < len(starts) else n
        if s < n:
            blocks.append((s, e))
    if not blocks:
        blocks = [(0, n)]
    return blocks


def build_orders_blocks(
    fechas: List[dt.date],
    esperado: List[float],
    stock_ini_day0: float,
    target_override: Dict[str, float],
    cfg: OrderCfg,
    cal: DeliveryCalendar,
) -> Tuple[List[int], List[float]]:
    """Compute per-day Pedido and Stock_fin using block logic.
    Pedido is placed only on the first day of each block (rounded once to multiple).
    """
    n = len(fechas)
    pedido: List[int] = [0] * n
    stock_fin: List[float] = [0.0] * n
    stock_ini = float(stock_ini_day0)
    E = [float(x or 0.0) for x in esperado]
    blocks = _compute_blocks(fechas, cal)
    for (s, e) in blocks:
        demanda_b = float(sum(E[s:e]))
        target_prox = _target_efectivo(e - 1, fechas, E, target_override, cfg)
        pedido_bruto_b = max(0.0, target_prox + demanda_b - stock_ini)
        pedido_b = _ceil_to_multiple(pedido_bruto_b, cfg.multiplo)
        if s < n:
            pedido[s] = int(pedido_b)
        for i in range(s, e):
            add = pedido[i] if i == s else 0
            sf = stock_ini + add - E[i]
            stock_fin[i] = float(sf)
            stock_ini = float(sf)
    return pedido, stock_fin


def build_orders_daily(
    fechas: List[dt.date],
    esperado: List[float],
    stock_ini_day0: float,
    target_override: Dict[str, float],
    cfg: OrderCfg,
) -> Tuple[List[int], List[float]]:
    """Daily mode (equivalent to blocks of size 1)."""
    n = len(fechas)
    pedido: List[int] = [0] * n
    stock_fin: List[float] = [0.0] * n
    stock_ini = float(stock_ini_day0)
    E = [float(x or 0.0) for x in esperado]
    for i in range(n):
        target_eff = _target_efectivo(i, fechas, E, target_override, cfg)
        pedido_bruto = max(0.0, target_eff + E[i] - stock_ini)
        p = _ceil_to_multiple(pedido_bruto, cfg.multiplo)
        pedido[i] = int(p)
        sf = stock_ini + p - E[i]
        stock_fin[i] = float(sf)
        stock_ini = float(sf)
    return pedido, stock_fin


def _read_delivery_calendar_from_excel(xls_bytes: bytes) -> DeliveryCalendar:
    """Read CFG_DELIVERY_WEEK from CFG!B8:H8 and ENTREGAS_EXC sheet if present."""
    week_bits = [1, 1, 1, 1, 1, 1, 1]
    exceptions: Dict[str, int] = {}
    try:
        xls = pd.ExcelFile(io.BytesIO(xls_bytes))
        if "CFG" in xls.sheet_names:
            df = xls.parse("CFG", header=None)
            # B8:H8 are row index 7 and columns 1..7
            try:
                row = df.iloc[7, 1:8]
                vals = []
                for v in row.tolist():
                    try:
                        vals.append(1 if int(pd.to_numeric(v, errors="coerce") or 0) == 1 else 0)
                    except Exception:
                        vals.append(1)
                if len(vals) == 7:
                    week_bits = vals
            except Exception:
                pass
        if "ENTREGAS_EXC" in xls.sheet_names:
            exc_df = xls.parse("ENTREGAS_EXC")
            cand = {str(c).strip().lower(): c for c in exc_df.columns}
            dcol = cand.get("fecha") or list(exc_df.columns)[0]
            ecol = cand.get("entrega") or (list(exc_df.columns)[1] if len(exc_df.columns) > 1 else None)
            if dcol is not None and ecol is not None:
                dd = ensure_datetime(exc_df[dcol]).dt.normalize()
                vv = pd.to_numeric(exc_df[ecol], errors="coerce").fillna(0).astype(int)
                for d, v in zip(dd, vv):
                    if pd.notna(d):
                        exceptions[pd.Timestamp(d).date().isoformat()] = 1 if int(v) == 1 else 0
    except Exception:
        pass
    return DeliveryCalendar(week_bits=week_bits, exceptions=exceptions)


def read_delivery_cfg(cfg: Dict, exceptions_rows: Optional[Iterable[Dict]] = None) -> DeliveryCalendar:
    """Build DeliveryCalendar from simple dict cfg and optional exceptions rows.
    cfg expects key 'CFG_DELIVERY_WEEK' as list of 7 ints [Mon..Sun].
    exceptions_rows: iterable of dicts with keys 'Fecha' and 'Entrega' (0/1),
    accepting dates in '%d/%m/%Y' or ISO 'YYYY-MM-DD'.
    """
    wb = cfg.get("CFG_DELIVERY_WEEK") if isinstance(cfg, dict) else None
    if isinstance(wb, str):
        try:
            wb = [int(x.strip()) for x in wb.split(",")]
        except Exception:
            wb = None
    if not (isinstance(wb, (list, tuple)) and len(wb) == 7):
        wb = [1, 1, 1, 1, 1, 1, 1]
    exc: Dict[str, int] = {}
    if exceptions_rows:
        for r in exceptions_rows:
            f = str(r.get("Fecha") or r.get("fecha") or "").strip()
            v = r.get("Entrega") if "Entrega" in r else r.get("entrega")
            if not f:
                continue
            d_iso = None
            for fmt in (DATE_FMT, "%Y-%m-%d"):
                try:
                    d = pd.to_datetime(f, format=fmt, errors="coerce")
                    if pd.notna(d):
                        d_iso = pd.Timestamp(d).date().isoformat()
                        break
                except Exception:
                    continue
            if d_iso is not None:
                try:
                    exc[d_iso] = 1 if int(v) == 1 else 0
                except Exception:
                    pass
    return DeliveryCalendar(week_bits=[1 if int(x) == 1 else 0 for x in wb], exceptions=exc)


def validate_strict_columns(df: pd.DataFrame, required_cols: List[str], sheet: str):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"No se detectó la columna {missing[0]} en la hoja {sheet}.")


def try_read_pedido_date(xls: pd.ExcelFile) -> Optional[pd.Timestamp]:
    try:
        if "PEDIDO_CP" not in xls.sheet_names:
            return None
        df = xls.parse("PEDIDO_CP", header=None, usecols=[3], nrows=2)
        if df.shape[0] >= 2:
            val = df.iloc[1, 0]
            if pd.isna(val):
                return None
            if isinstance(val, (int, float)):
                base = pd.Timestamp("1899-12-30")
                return (base + pd.to_timedelta(int(val), unit="D")).normalize()
            if isinstance(val, str):
                dt1 = pd.to_datetime(val.strip(), format=DATE_FMT, errors="coerce")
                return pd.Timestamp(dt1).normalize() if pd.notna(dt1) else None
            else:
                return None
    except Exception:
        return None
    return None


# ---------------- Calculations ----------------
def build_transitions_from_ventas(
    ventas: pd.DataFrame,
    target_date: pd.Timestamp,
    window_days: int,
    use_dow: bool,
    use_wom: bool,
    alpha: float = 1.0,
) -> Tuple[pd.DataFrame, np.ndarray]:
    dcol, scol = ES_VENTAS["date"], ES_VENTAS["state"]
    v = ventas[[dcol, scol]].dropna().copy()
    v[dcol] = ensure_datetime(v[dcol])
    if scol in v.columns:
        v[scol] = normalize_state_series(v[scol])
    v = v.dropna(subset=[dcol]).sort_values(dcol)
    min_ts = target_date - pd.Timedelta(days=int(window_days)) if window_days and window_days > 0 else pd.Timestamp.min
    v = v[(v[dcol] < target_date) & (v[dcol] >= min_ts)]
    v["estado_hoy"] = v[scol].astype(str)
    v["estado_manana"] = v[scol].shift(-1).astype(str)
    v["fecha_manana"] = v[dcol].shift(-1)
    v = v.dropna(subset=["estado_hoy", "estado_manana", "fecha_manana"])
    if use_dow:
        v = v[v["fecha_manana"].dt.weekday == int(target_date.weekday())]
    if use_wom:
        target_wom = week_of_month(target_date)
        v = v[((v["fecha_manana"].dt.day - 1) // 7 + 1) == target_wom]
    idx = {s: i for i, s in enumerate(STATES)}
    counts = np.zeros((len(STATES), len(STATES)), dtype=float)
    for _, r in v.iterrows():
        i = idx.get(r["estado_hoy"])
        j = idx.get(r["estado_manana"])
        if i is None or j is None:
            continue
        counts[i, j] += 1.0
    m = counts + alpha
    row_sums = m.sum(axis=1, keepdims=True)
    with np.errstate(invalid="ignore"):
        probs = np.divide(m, row_sums, out=np.zeros_like(m), where=row_sums != 0)
    for r in range(probs.shape[0]):
        if row_sums[r, 0] == 0:
            probs[r, :] = 1.0 / len(STATES)
    return pd.DataFrame(probs, index=STATES, columns=STATES), counts


def compute_Q(diario: pd.DataFrame, map_d: Dict[str, str], target_date: pd.Timestamp, window_days: int) -> pd.DataFrame:
    dcol, icol, scol, ucol = map_d["date"], map_d["item"], map_d["state"], map_d["units"]
    if not ucol or ucol not in diario.columns:
        return pd.DataFrame()
    d = diario[[dcol, icol, scol, ucol]].copy()
    d[dcol] = ensure_datetime(d[dcol])
    d[ucol] = pd.to_numeric(d[ucol], errors="coerce").fillna(0.0)
    d = d.dropna(subset=[dcol, icol, scol])
    if scol in d.columns:
        d[scol] = normalize_state_series(d[scol])
    # Normalize item names for robust matching
    d["__item_key__"] = d[icol].map(normalize_item_name)
    min_ts = target_date - pd.Timedelta(days=int(window_days)) if window_days and window_days > 0 else d[dcol].min()
    d = d[(d[dcol] < target_date) & (d[dcol] >= min_ts)]
    d["__state__"] = d[scol].astype(str)
    daily = d.groupby([dcol, "__item_key__", "__state__"], dropna=False)[ucol].sum().reset_index()
    Q = daily.groupby(["__item_key__", "__state__"], dropna=False)[ucol].mean().unstack(fill_value=0.0)
    for s in STATES:
        if s not in Q.columns:
            Q[s] = 0.0
    return Q[STATES].clip(lower=0.0)


def read_pred_week(xls_bytes: bytes) -> List[Tuple[pd.Timestamp, str]]:
    try:
        xls = pd.ExcelFile(io.BytesIO(xls_bytes))
        if "PRED_VENTA" not in xls.sheet_names:
            return []
        df = xls.parse("PRED_VENTA", header=None)
        fechas = parse_fecha_dmY(df.iloc[2:13, 1])
        estados = normalize_state_series(df.iloc[2:13, 8])
        out: List[Tuple[pd.Timestamp, str]] = []
        for d, s in zip(fechas, estados):
            if pd.notna(d) and isinstance(s, str) and s.strip():
                out.append((pd.Timestamp(d).normalize(), s.strip()))
        return out
    except Exception:
        return []


def build_weekly_grid(Q: pd.DataFrame, xls_bytes: bytes, stock_df: pd.DataFrame) -> pd.DataFrame:
    try:
        xls = pd.ExcelFile(io.BytesIO(xls_bytes))
        dfp = xls.parse("PEDIDO_CP") if "PEDIDO_CP" in xls.sheet_names else pd.DataFrame()
    except Exception:
        dfp = pd.DataFrame()
    if not dfp.empty and all(c in dfp.columns for c in ["RUBRO", "ITEM"]):
        items = dfp[["RUBRO", "ITEM"]].dropna().copy()
        items.columns = ["Rubro", "Item"]
    else:
        s = stock_df.copy()
        if s.empty:
            return pd.DataFrame()
        if "rubro" not in s.columns:
            s["rubro"] = ""
        items = s[["rubro", "item"]].drop_duplicates().copy()
        items.columns = ["Rubro", "Item"]
    week = read_pred_week(xls_bytes)
    if not week:
        return items
    grid = items.copy()
    for dt, st_name in week:
        col = dt.strftime("%d/%m/%Y")
        vals = []
        for _, row in items.iterrows():
            item = str(row["Item"]) 
            key = normalize_item_name(item)
            v = float(pd.to_numeric(Q.loc[key, st_name], errors="coerce")) if (not Q.empty and key in Q.index and st_name in Q.columns) else 0.0
            vals.append(v)
        grid[col] = vals
    return grid


def build_weekly_orders(
    Q: pd.DataFrame,
    xls_bytes: bytes,
    stock_df: pd.DataFrame,
    order_multiple: int = 1,
    target: float = 0.0,
    enable_am: bool = True,
    am_frac: float = 0.5,
) -> pd.DataFrame:
    try:
        xls = pd.ExcelFile(io.BytesIO(xls_bytes))
        dfp = xls.parse("PEDIDO_CP") if "PEDIDO_CP" in xls.sheet_names else pd.DataFrame()
    except Exception:
        dfp = pd.DataFrame()
    if not dfp.empty and all(c in dfp.columns for c in ["RUBRO", "ITEM"]):
        items = dfp[["RUBRO", "ITEM"]].dropna().copy()
        items.columns = ["Rubro", "Item"]
    else:
        s = stock_df.copy()
        if s.empty:
            return pd.DataFrame()
        if "rubro" not in s.columns:
            s["rubro"] = ""
        items = s[["rubro", "item"]].drop_duplicates().copy()
        items.columns = ["Rubro", "Item"]
    week = read_pred_week(xls_bytes)
    if not week:
        return pd.DataFrame()
    dates = [pd.Timestamp(dt).normalize() for dt, _ in week]
    # stock por fecha (mapa normalizado)
    stock_by_date: Dict[pd.Timestamp, Dict[str, float]] = {}
    if not stock_df.empty and "fecha" in stock_df.columns:
        s = stock_df.copy()
        s["fecha"] = ensure_datetime(s["fecha"]).dt.normalize()
        g = s.groupby(["fecha", "item"], dropna=False)["stock"].sum().reset_index()
        for d, g_d in g.groupby("fecha"):
            stock_by_date[pd.Timestamp(d)] = {normalize_item_name(r["item"]): float(pd.to_numeric(r["stock"], errors="coerce") or 0.0) for _, r in g_d.iterrows()}
    else:
        # estático (igual para todos los días)
        static_map: Dict[str, float] = {}
        if not stock_df.empty:
            for _, r in stock_df.iterrows():
                static_map[normalize_item_name(r.get("item", ""))] = float(pd.to_numeric(r.get("stock", 0.0), errors="coerce") or 0.0)
        for dt, _ in week:
            stock_by_date[pd.Timestamp(dt)] = static_map
    # preparar override opcional por día (TARGET_DIA! A:Fecha, B:Target)
    target_override: Dict[pd.Timestamp, float] = {}
    try:
        xls_o = pd.ExcelFile(io.BytesIO(xls_bytes))
        if "TARGET_DIA" in xls_o.sheet_names:
            df_t = xls_o.parse("TARGET_DIA")
            # detect columns
            cand = {str(c).strip().lower(): c for c in df_t.columns}
            dcol = cand.get("fecha") or list(df_t.columns)[0]
            tcol = cand.get("target") or list(df_t.columns)[1]
            dd = ensure_datetime(df_t[dcol]).dt.normalize()
            tv = pd.to_numeric(df_t[tcol], errors="coerce")
            for d, v in zip(dd, tv):
                if pd.notna(d) and pd.notna(v):
                    target_override[pd.Timestamp(d)] = float(v)
    except Exception:
        target_override = {}

    # preparar grilla y columnas de fechas
    grid = items.copy()
    for dt in dates:
        col = pd.Timestamp(dt).strftime("%d/%m/%Y")
        if col not in grid.columns:
            grid[col] = 0
    # Nueva lógica: pedidos semanales por bloques basados en calendario CFG/ENTREGAS_EXC
    am_frac = float(min(1.0, max(0.0, am_frac)))
    order_multiple = int(max(1, order_multiple))
    target = float(max(0.0, target))
    cal = _read_delivery_calendar_from_excel(xls_bytes)
    use_blocks = (0 in cal.week_bits) or (len(cal.exceptions) > 0)
    for ridx, row in items.iterrows():
        item = str(row["Item"]) 
        key = normalize_item_name(item)
        stock_ini = float(stock_by_date.get(dates[0], {}).get(key, 0.0))
        expected_seq: List[float] = []
        for dt_i, st_name in week:
            expected_seq.append(
                float(pd.to_numeric(Q.loc[key, st_name], errors="coerce")) if (not Q.empty and key in Q.index and st_name in Q.columns) else 0.0
            )
        t_override_map: Dict[str, float] = {}
        for d in dates:
            if d in target_override:
                t_override_map[pd.Timestamp(d).date().isoformat()] = float(target_override[d])
        cfg = OrderCfg(
            target_global=target,
            multiplo=order_multiple,
            enable_am=enable_am,
            am_frac=am_frac,
            enable_blocks=None,
        )
        fechas_seq = [pd.Timestamp(d).date() for d in dates]
        if use_blocks:
            pedidos_seq, _ = build_orders_blocks(fechas_seq, expected_seq, stock_ini, t_override_map, cfg, cal)
        else:
            pedidos_seq, _ = build_orders_daily(fechas_seq, expected_seq, stock_ini, t_override_map, cfg)
        for i, dt_val in enumerate(dates):
            col = dt_val.strftime("%d/%m/%Y")
            grid.at[ridx, col] = int(max(0, pedidos_seq[i]))
    return grid
    # helper redondeo a múltiplos
    def _ceil_multiple(x: float, m: int) -> int:
        if m and m > 1:
            return int(np.ceil(max(0.0, x) / m) * m)
        return int(np.ceil(max(0.0, x)))
    # calcular por ítem con arrastre
    # clamp params
    am_frac = float(min(1.0, max(0.0, am_frac)))
    order_multiple = int(max(1, order_multiple))
    target = float(max(0.0, target))

    for ridx, row in items.iterrows():
        item = str(row["Item"]) 
        key = normalize_item_name(item)
        stock_ini = float(stock_by_date.get(dates[0], {}).get(key, 0.0))
        pedidos_seq: List[int] = []
        s_cur = stock_ini
        for i, (dt, st_name) in enumerate(week):
            dt = pd.Timestamp(dt).normalize()
            expected = float(pd.to_numeric(Q.loc[key, st_name], errors="coerce")) if (not Q.empty and key in Q.index and st_name in Q.columns) else 0.0
            # Target efectivo por día: override -> max(CFG_TARGET, auto)
            t_override = target_override.get(dt)
            auto = 0.0
            if enable_am:
                # próximo día si existe
                if i + 1 < len(week):
                    _, st_next = week[i + 1]
                    auto = am_frac * (float(pd.to_numeric(Q.loc[key, st_next], errors="coerce")) if (not Q.empty and key in Q.index and st_next in Q.columns) else 0.0)
                else:
                    auto = 0.0
            t_eff = float(t_override) if (t_override is not None and not pd.isna(t_override)) else max(target, auto)
            pedido = _ceil_multiple(max(0.0, t_eff + expected - s_cur), order_multiple)
            s_fin = s_cur + pedido - expected
            pedidos_seq.append(int(pedido))
            s_cur = float(s_fin)
        # asignar a columnas
        for i, dt in enumerate(dates):
            col = dt.strftime("%d/%m/%Y")
            grid.at[ridx, col] = int(pedidos_seq[i])
    return grid


def _weekday_abbrev(ts: pd.Timestamp) -> str:
    # LUN=0..DOM=6
    names = {0: "LUN", 1: "MAR", 2: "MIE", 3: "JUE", 4: "VIE", 5: "SAB", 6: "DOM"}
    try:
        return names[int(pd.Timestamp(ts).weekday())]
    except Exception:
        return ""


def build_operational_table(weekly_orders: pd.DataFrame, stock_df: pd.DataFrame, xls_bytes: bytes) -> pd.DataFrame:
    # Espera weekly_orders con columnas: Rubro, Item, y fechas dd/mm/yyyy
    if weekly_orders is None or weekly_orders.empty:
        return pd.DataFrame(columns=["RUBRO", "ITEM", "STOCK (día_0)", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"])

    # Identificar columnas de fechas en weekly_orders
    base_cols = ["Rubro", "Item"]
    date_cols = [c for c in weekly_orders.columns if c not in base_cols]
    # Parsear a datetime
    parsed_dates = []
    for c in date_cols:
        dt = pd.to_datetime(c, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            parsed_dates.append(pd.Timestamp(dt).normalize())
    if not parsed_dates:
        return pd.DataFrame(columns=["RUBRO", "ITEM", "STOCK (día_0)", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"])

    # Día 0: preferir PRED_VENTA!B3 si existe; si no, mínima fecha de la semana
    day0 = min(parsed_dates)
    try:
        xls0 = pd.ExcelFile(io.BytesIO(xls_bytes))
        if "PRED_VENTA" in xls0.sheet_names:
            pv = xls0.parse("PRED_VENTA", header=None)
            if pv.shape[0] > 2 and pv.shape[1] > 1:
                b3 = pd.to_datetime(str(pv.iloc[2, 1]).strip(), format=DATE_FMT, errors="coerce")
                if pd.notna(b3):
                    day0 = pd.Timestamp(b3).normalize()
    except Exception:
        pass

    # Stock (día_0): preferir HOJA "INVENTARIO P PEDIDO" (A=ITEM, B=CANTIDAD). Fallback a PEDIDO_CP Col C
    stock_map_day0: Dict[str, float] = {}
    try:
        xls = pd.ExcelFile(io.BytesIO(xls_bytes))
        # 1) Inventario por pedido: columnas A (item), B (cantidad)
        if "INVENTARIO P PEDIDO" in xls.sheet_names:
            inv = xls.parse("INVENTARIO P PEDIDO", header=None)
            if not inv.empty and inv.shape[1] >= 2:
                items_ser = inv.iloc[:, 0].astype(str)
                qty_ser = pd.to_numeric(inv.iloc[:, 1], errors="coerce").fillna(0.0)
                for it, sv in zip(items_ser, qty_ser):
                    stock_map_day0[str(it)] = float(sv)
        # 2) Si no hay hoja de inventario, intentar PEDIDO_CP columna C
        if not stock_map_day0 and "PEDIDO_CP" in xls.sheet_names:
            dfp = xls.parse("PEDIDO_CP")
            if not dfp.empty and dfp.shape[1] >= 3:
                item_col = "ITEM" if "ITEM" in dfp.columns else dfp.columns[1]
                stock_col = dfp.columns[2]
                items_ser = dfp[item_col].astype(str)
                stock_ser = pd.to_numeric(dfp[stock_col], errors="coerce").fillna(0.0)
                for it, sv in zip(items_ser, stock_ser):
                    stock_map_day0[str(it)] = float(sv)
    except Exception:
        stock_map_day0 = {}
    # Fallback al stock_df si no se pudo leer PEDIDO_CP
    if not stock_map_day0:
        if stock_df is not None and not stock_df.empty and "fecha" in stock_df.columns:
            s = stock_df.copy()
            s["fecha"] = ensure_datetime(s["fecha"]).dt.normalize()
            day0_rows = s.loc[s["fecha"] == day0]
            for _, r in day0_rows.iterrows():
                stock_map_day0[str(r.get("item", ""))] = float(pd.to_numeric(r.get("stock", 0.0), errors="coerce") or 0.0)
        else:
            if stock_df is not None and not stock_df.empty:
                for _, r in stock_df.iterrows():
                    stock_map_day0[str(r.get("item", ""))] = float(pd.to_numeric(r.get("stock", 0.0), errors="coerce") or 0.0)

    # Mapear fechas -> abreviaturas
    desired_order = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]
    rows = []
    for _, r in weekly_orders.iterrows():
        rubro = r.get("Rubro", "")
        item = r.get("Item", "")
        # Stock día_0: clamp a >= 0 y marcar alerta si el origen fue negativo
        raw_stock = float(stock_map_day0.get(str(item), 0.0))
        alert = "" if raw_stock >= 0 else "AMARILLO"
        shown_stock = int(max(0, raw_stock))
        rec = {"RUBRO": rubro, "ITEM": item, "STOCK (día_0)": shown_stock, "ALERTA_INV": alert}
        for dname in desired_order:
            rec[dname] = 0
        for c in date_cols:
            dt = pd.to_datetime(c, format=DATE_FMT, errors="coerce")
            if pd.notna(dt):
                ab = _weekday_abbrev(pd.Timestamp(dt))
                if ab in desired_order:
                    rec[ab] = int(max(0, pd.to_numeric(r.get(c, 0), errors="coerce") or 0))
        rows.append(rec)

    df_out = pd.DataFrame(rows, columns=["RUBRO", "ITEM", "STOCK (día_0)", "ALERTA_INV"] + desired_order)
    return df_out


def detect_and_prepare_strict(
    xls_bytes: bytes,
    alpha: float,
    target_date: pd.Timestamp,
    window_days: int,
    use_dow: bool,
    use_wom: bool,
):
    xls = pd.ExcelFile(io.BytesIO(xls_bytes))
    for sheet in ["VENTA_DIARIA", "DIARIO", "PEDIDO_CP"]:
        if sheet not in xls.sheet_names:
            raise ValueError(f"Falta la hoja {sheet} en el Excel.")
    ventas = pd.read_excel(io.BytesIO(xls_bytes), sheet_name="VENTA_DIARIA")
    diario = pd.read_excel(io.BytesIO(xls_bytes), sheet_name="DIARIO")
    stock = pd.read_excel(io.BytesIO(xls_bytes), sheet_name="PEDIDO_CP")
    ventas, diario, stock = map(sanitize_for_display, [ventas, diario, stock])
    # Estandariza encabezados
    ventas = std_cols(ventas)
    diario = std_cols(diario)
    stock = std_cols(stock)
    # Asegura fechas en VENTA_DIARIA
    if ES_VENTAS["date"] in ventas.columns:
        ventas[ES_VENTAS["date"]] = ensure_datetime(ventas[ES_VENTAS["date"]])
    validate_strict_columns(ventas, [ES_VENTAS["date"], ES_VENTAS["state"]], "VENTA_DIARIA")
    validate_strict_columns(diario, [ES_DIARIO["date"], ES_DIARIO["item"], ES_DIARIO["state"], ES_DIARIO["units"]], "DIARIO")
    maps = {"diario": ES_DIARIO.copy()}
    probs_used, counts_used = build_transitions_from_ventas(ventas, target_date, window_days, use_dow, use_wom, alpha=alpha)
    Q = compute_Q(diario, maps["diario"], target_date, window_days)
    # stock largo
    stock_df = pd.DataFrame()
    for cols in [("Fecha", "ITEM", "Stock"), ("Fecha", "ITEM", "Disponible"), ("Fecha", "ITEM", "Saldo")]:
        if all(c in stock.columns for c in cols):
            s = stock[list(cols)].copy()
            rubro_col = "RUBRO" if "RUBRO" in stock.columns else None
            if rubro_col:
                s[rubro_col] = stock[rubro_col]
            s.columns = ["fecha", "item", "stock"] + (["rubro"] if rubro_col else [])
            s["fecha"] = ensure_datetime(s["fecha"]) 
            stock_df = s.copy()
            break
    # stock ancho
    if stock_df.empty and "ITEM" in stock.columns:
        id_vars = ["ITEM"] + (["RUBRO"] if "RUBRO" in stock.columns else [])
        other = [c for c in stock.columns if c not in id_vars]
        date_cols = []
        for c in other:
            dt = pd.to_datetime(c, format=DATE_FMT, errors="coerce")
            if pd.notna(dt):
                date_cols.append(c)
        if date_cols:
            parsed = parse_fecha_robusta(pd.Series(date_cols)).dt.normalize()
            if not any(parsed == target_date.normalize()):
                raise ValueError(f"No hay columna de stock para la fecha {target_date.strftime('%d/%m/%Y')} en PEDIDO_CP.")
            melted = stock[id_vars + date_cols].melt(id_vars=id_vars, var_name="fecha", value_name="stock")
            melted["fecha"] = ensure_datetime(melted["fecha"]) 
            melted.rename(columns={"ITEM": "item", "RUBRO": "rubro"}, inplace=True)
            stock_df = melted.dropna(subset=["fecha"]).copy()
    # stock estatico (tercera columna): permite que ITEM/RUBRO sean columnas sin encabezado
    if stock_df.empty and len(stock.columns) >= 3:
        item_col = "ITEM" if "ITEM" in stock.columns else stock.columns[1]
        rubro_col = "RUBRO" if "RUBRO" in stock.columns else stock.columns[0]
        third = stock.columns[2]
        try:
            s = stock[[item_col, third] + ([rubro_col] if rubro_col else [])].copy()
        except Exception:
            s = stock[[item_col, third]].copy()
            rubro_col = None
        ren = {item_col: "item", third: "stock"}
        if rubro_col:
            ren[rubro_col] = "rubro"
        s.rename(columns=ren, inplace=True)
        # limpiar posibles filas de encabezado repetidas
        s["item"] = s["item"].astype(str).str.strip()
        s = s[~s["item"].str.upper().isin(["", "ITEM", "RUBRO", "STOCK"])]
        s["stock"] = pd.to_numeric(s["stock"], errors="coerce").fillna(0.0)
        stock_df = s
    return ventas, diario, maps, probs_used, counts_used, stock_df, Q


def yesterday_state(ventas: pd.DataFrame, target_date: pd.Timestamp) -> Tuple[str, bool]:
    dcol, scol = ES_VENTAS["date"], ES_VENTAS["state"]
    v = ventas[[dcol, scol]].copy()
    v[dcol] = ensure_datetime(v[dcol])
    yday = target_date - pd.Timedelta(days=1)
    row = v.loc[v[dcol].dt.normalize() == yday.normalize()]
    if not row.empty:
        return str(row.iloc[0][scol]), True
    return "— OK", False


def predict_for_date(
    ventas: pd.DataFrame,
    diario: pd.DataFrame,
    maps: Dict[str, Dict[str, Optional[str]]],
    probs: pd.DataFrame,
    Q: pd.DataFrame,
    stock_df: pd.DataFrame,
    target_date: pd.Timestamp,
    state_yesterday: str,
    order_multiple: int = 1,
) -> pd.DataFrame:
    s_for_date = pd.DataFrame(columns=["rubro", "item", "stock"])
    if not stock_df.empty:
        s = stock_df.copy()
        if "fecha" in s.columns:
            s["fecha"] = ensure_datetime(s["fecha"]) 
            s_day = s[s["fecha"].dt.normalize() == target_date.normalize()]
        else:
            s_day = s
        if "rubro" not in s_day.columns:
            s_day["rubro"] = ""
        s_for_date = s_day[["rubro", "item", "stock"]].copy()
        s_for_date["stock"] = pd.to_numeric(s_for_date["stock"], errors="coerce").fillna(0.0)
    if Q.empty:
        Q_use = pd.DataFrame(0.0, index=[normalize_item_name(x) for x in s_for_date.get("item", pd.Series([], dtype=str)).astype(str).unique()], columns=STATES)
    else:
        Q_use = Q.copy()
        for s in STATES:
            if s not in Q_use.columns:
                Q_use[s] = 0.0
        Q_use = Q_use[STATES]
    s_y = state_yesterday if state_yesterday in probs.index else "- OK"
    p_row = probs.loc[s_y] if s_y in probs.index else pd.Series(np.full((len(STATES),), 1.0 / len(STATES)), index=STATES)
    rows = []
    for _, rec in s_for_date.iterrows():
        rubro = rec.get("rubro", "")
        item = str(rec["item"]) 
        key = normalize_item_name(item)
        expected = float(pd.to_numeric(Q_use.loc[key, :], errors="coerce").reindex(STATES).fillna(0.0).values.dot(p_row.values)) if key in Q_use.index else 0.0
        stock_val = float(np.nan_to_num(rec.get("stock", 0.0)))
        diff = max(0.0, expected - stock_val)
        pedido = int(np.ceil(diff / order_multiple) * order_multiple) if order_multiple and order_multiple > 1 else int(np.ceil(diff))
        rows.append({"Rubro": rubro, "Item": item, "Esperado (u)": float(expected), "Stock (u)": float(stock_val), "Pedido sugerido (u)": int(pedido)})
    return pd.DataFrame(rows)


# ---------------- UI ----------------
st.set_page_config(page_title="MADAME CEPHALPOD ORACLE", layout="wide")


def _b64_image(path: str) -> Optional[str]:
    p = Path(path)
    if not p.exists():
        return None
    try:
        data = p.read_bytes()
        return base64.b64encode(data).decode()
    except Exception:
        return None


_bg_candidates = [
    "FONDO WEB.jpeg",
    "FONDO WEB.jpg",
    "FONDO_WEB.jpeg",
    "FONFO WEB.jpeg",
]
_bg_b64 = None
for _fname in _bg_candidates:
    _bg_b64 = _b64_image(_fname)
    if _bg_b64:
        break
if _bg_b64:
    st.markdown(
        f"""
        <style>
        .stApp {{
            background: url(data:image/jpeg;base64,{_bg_b64}) no-repeat center fixed;
            background-size: cover;
        }}
        section[data-testid="stSidebar"] > div {{ background-color: rgba(0,0,0,0.55); }}
        .block-container {{ background-color: rgba(0,0,0,0.35); padding: 1rem 1.5rem; border-radius: 8px; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

st.markdown("<h1 style='text-align:center'>MADAME CEPHALPOD ORACLE</h1>", unsafe_allow_html=True)
st.caption("App local (offline). Excel con hojas VENTA_DIARIA, DIARIO y PEDIDO_CP.")

with st.sidebar:
    st.header("Entrada")
    st.caption("Usando archivo local")
    st.code("PITONISA.xlsx")
    alpha = st.number_input("Suavizado de Laplace (alpha)", min_value=0.0, value=1.0, step=0.5)
    st.subheader("Parametros")
    use_dow = st.toggle("Condicionar por dia de semana", value=True)
    use_wom = st.toggle("Condicionar por semana del mes", value=True)
    window_days = st.number_input("Ventana (dias hacia atras)", min_value=0, value=120, step=7, help="0 = usar todo el historial")
    # Intentar leer defaults desde CFG del Excel local
    _p = Path(__file__).parent / "PITONISA.xlsx"
    _t_def, _m_def, _am_on, _am_frac, _item_filter_def = (0.0, 1, True, 0.5, "")
    try:
        if _p.exists():
            _t_def, _m_def, _am_on, _am_frac, _item_filter_def = read_cfg_defaults(_read_bytes_retry(_p))
    except Exception:
        _t_def, _m_def, _am_on, _am_frac, _item_filter_def = (0.0, 1, True, 0.5, "")
    target_stock = st.number_input(
        "Target de stock al cierre",
        min_value=0.0,
        value=float(_t_def),
        step=1.0,
        help="Stock objetivo al cierre del día (0 = terminar justo).",
    )
    order_multiple = st.number_input(
        "Multiplo de pedido",
        min_value=1,
        value=int(_m_def),
        step=1,
        help="Tamaño de caja/pack (≥1); se redondea hacia arriba a este múltiplo.",
    )
    enable_am = st.toggle("Cobertura mañana (AM)", value=bool(_am_on), help="Si está activo, cubre una fracción del esperado del día siguiente como stock objetivo.")
    am_frac = st.number_input("Fracción AM (0-1)", min_value=0.0, max_value=1.0, value=float(_am_frac), step=0.05)
    item_filter = st.text_input("Filtro por ítem (opcional)", value=str(_item_filter_def or ""))
    # Calendario semanal de entregas (CFG!B8:H8)
    week_bits_default = [1, 1, 1, 1, 1, 1, 1]
    try:
        if _p.exists():
            _cal = _read_delivery_calendar_from_excel(_read_bytes_retry(_p))
            if _cal and _cal.week_bits and len(_cal.week_bits) == 7:
                week_bits_default = [1 if int(v) == 1 else 0 for v in _cal.week_bits]
    except Exception:
        week_bits_default = [1, 1, 1, 1, 1, 1, 1]
    with st.form("cfg_form", clear_on_submit=False):
        st.subheader("Calendario semanal de entregas")
        st.caption("1 = hay entrega; 0 = sin entrega")
        days_names = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        cols_week = st.columns(7)
        week_bits_selected: List[int] = []
        for i, name in enumerate(days_names):
            with cols_week[i]:
                val = st.toggle(name, value=bool(week_bits_default[i]), key=f"dow_{i}")
                week_bits_selected.append(1 if val else 0)
        colA, colB = st.columns(2)
        save = colA.form_submit_button("Guardar CFG")
        reset = colB.form_submit_button("Reset CFG")
        if save:
            ok = write_cfg_to_file(
                _p,
                target_stock,
                order_multiple,
                enable_am=enable_am,
                am_frac=am_frac,
                item_filter=item_filter,
                delivery_week=week_bits_selected,
            )
            st.success("CFG guardado" if ok else "No se pudo guardar CFG")
            # Evitar recálculo inmediato tras guardar; requerir botón explícito
            st.session_state["do_recalc"] = False
            st.info("Cambios guardados. Pulsa 'Aplicar y recalcular'.")
        if reset:
            ok = write_cfg_to_file(
                _p,
                0.0,
                1,
                enable_am=True,
                am_frac=0.5,
                item_filter="",
                delivery_week=[1, 1, 1, 1, 1, 1, 1],
            )
            st.success("CFG reseteado" if ok else "No se pudo guardar CFG")
            st.session_state["do_recalc"] = False
            st.info("CFG por defecto cargado. Pulsa 'Aplicar y recalcular'.")

    # Botón explícito para recalcular
    if st.button("Aplicar y recalcular", type="primary"):
        st.session_state["do_recalc"] = True
    btn_run = st.button("Calcular prediccion")
    st.divider()
    st.subheader("CSV (opcional)")
    st.caption("Importa un CSV (Fecha, Item, Esperado, Stock inicial opcional).")
    csv_file = st.file_uploader("Subir CSV", type=["csv"]) 
    enable_blocks = st.toggle("Modo bloques sin retiro (usa columna 'Retira?')", value=False)

    # Notas / Especificación (sidebar)
    def load_notes_md(path: str = "docs/NOTAS_PITONISA.md") -> str:
        p = Path(path)
        try:
            return p.read_text(encoding="utf-8") if p.exists() else "# (Sin notas)\n"
        except Exception:
            return "# (Sin notas)\n"

    with st.sidebar.expander(" Notas / Especificación"):
        st.markdown(load_notes_md(), unsafe_allow_html=False)
        q = st.text_input("Buscar en notas", key="notes_query")
        if q:
            try:
                import re as _re
                md = load_notes_md()
                hits = [l for l in md.splitlines() if _re.search(_re.escape(q), l, _re.IGNORECASE)]
                if hits:
                    st.write("Coincidencias:")
                    for h in hits[:20]:
                        st.write("- ", h[:120])
                else:
                    st.caption("Sin coincidencias.")
            except Exception:
                st.caption("(Error en búsqueda)")


xlsx_path = Path(__file__).parent / "PITONISA.xlsx"
if not xlsx_path.exists():
    st.error(f"No se encontro el archivo {xlsx_path.name} en la carpeta de la app.")
    st.stop()

st.success(f"Excel local detectado: {xlsx_path.name}")
try:
    # Estado de recálculo
    if "do_recalc" not in st.session_state:
        st.session_state["do_recalc"] = True  # primera carga

    file_bytes = _read_bytes_retry(xlsx_path)
    _xls = pd.ExcelFile(io.BytesIO(file_bytes))
    d2 = try_read_pedido_date(_xls)
    default_date = (d2.date() if d2 is not None else (pd.Timestamp.today() + pd.Timedelta(days=1)).date())
    target_date = st.date_input("Fecha objetivo", value=default_date)
    target_ts = pd.Timestamp(target_date)

    weekly_grid = pd.DataFrame(); weekly_orders = pd.DataFrame(); probs_used = pd.DataFrame(); ventas_df = pd.DataFrame(); diario_df = pd.DataFrame(); maps = {}; stock_df = pd.DataFrame(); Q = pd.DataFrame()
    if st.session_state.get("do_recalc", True):
        with st.spinner("Recalculando predicción y pedidos..."):
            ventas_df, diario_df, maps, probs_used, counts_used, stock_df, Q = detect_and_prepare_strict(
                file_bytes, alpha, target_ts, window_days, use_dow, use_wom
            )
            st.subheader("Matriz de probabilidades")
            st.caption("Orden de estados fijo: " + ", ".join(STATES))
            st.dataframe(sanitize_for_display(probs_used.round(2)))

            try:
                xls_view = pd.ExcelFile(io.BytesIO(file_bytes))
                if "PRED_VENTA" in xls_view.sheet_names:
                    pv = xls_view.parse("PRED_VENTA", header=None)
                    pv_view = pv.iloc[1:13, 0:9]
                    st.subheader("PREDICCION DE VENTA SEMANAL")
                    st.dataframe(sanitize_for_display(pv_view))
            except Exception:
                pass

            weekly_grid = build_weekly_grid(Q, file_bytes, stock_df)
            weekly_orders = build_weekly_orders(
                Q,
                file_bytes,
                stock_df,
                order_multiple=order_multiple,
                target=target_stock,
                enable_am=enable_am,
                am_frac=am_frac,
            )
        # apagar la bandera tras recálculo
        st.session_state["do_recalc"] = False
    else:
        st.info("Cambios guardados. Pulsa 'Aplicar y recalcular' para actualizar resultados.")
    # Aplicar filtro por ítem si corresponde
    try:
        if item_filter:
            patt = str(item_filter).strip().lower()
            if patt:
                if not weekly_grid.empty and "Item" in weekly_grid.columns:
                    weekly_grid = weekly_grid[weekly_grid["Item"].astype(str).str.lower().str.contains(patt, na=False)]
                if not weekly_orders.empty and "Item" in weekly_orders.columns:
                    weekly_orders = weekly_orders[weekly_orders["Item"].astype(str).str.lower().str.contains(patt, na=False)]
    except Exception:
        pass
    if csv_file is not None:
        try:
            df_csv_raw, csv_logs = read_csv_robust(csv_file.read())
            df_csv_map, map_used, map_logs = map_csv_columns(df_csv_raw)
            tovr = None
            try:
                xls_tmp = pd.ExcelFile(io.BytesIO(file_bytes))
                tovr = xls_tmp.parse("TARGET_DIA") if "TARGET_DIA" in xls_tmp.sheet_names else None
            except Exception:
                tovr = None
            pedidos_df, calc_logs = build_orders_from_csv(
                df_csv_map,
                target_stock,
                order_multiple,
                enable_am,
                am_frac,
                target_overrides=tovr,
                enable_blocks=enable_blocks,
            )
            # Filtro por ítem
            if item_filter:
                patt = str(item_filter).strip().lower()
                if patt:
                    pedidos_df = pedidos_df[pedidos_df["Item"].astype(str).str.lower().str.contains(patt, na=False)]
            st.subheader("PEDIDOS (desde CSV)")
            st.dataframe(sanitize_for_display(pedidos_df))
            # Pedido semanal dinámico
            sem = pedidos_df.groupby(["Fecha", "Item"], dropna=False)["Pedido"].sum().reset_index()
            if not sem.empty:
                st.subheader("Pedido sugerido semanal (unidades) [CSV]")
                st.dataframe(sanitize_for_display(sem.pivot_table(index=["Item"], columns="Fecha", values="Pedido", aggfunc="sum").reset_index()))
            # Logs
            if csv_logs or map_logs or calc_logs:
                st.caption("LOG importador/normalización:")
                for m in csv_logs + map_logs + calc_logs:
                    st.info(m)
        except Exception as e:
            st.error(f"Error al procesar CSV: {e}")

    if not weekly_grid.empty:
        st.subheader("Prediccion semanal por item (unidades)")
        st.dataframe(sanitize_for_display(weekly_grid))
    if not weekly_orders.empty:
        st.subheader("Pedido sugerido semanal (unidades)")
        st.dataframe(sanitize_for_display(weekly_orders))
    else:
        st.info("No hay pedido semanal para mostrar (revisa PRED_VENTA y que los items de PEDIDO_CP coincidan con DIARIO).")
    if Q.empty and weekly_grid.empty:
        st.info("No se pudo construir la prediccion semanal (Q vacio y/o sin PRED_VENTA).")

    # Tabla operativa (respuesta de CODEX)
    if not weekly_orders.empty:
        op_table = build_operational_table(weekly_orders, stock_df, file_bytes)
        if not op_table.empty:
            st.subheader("Tabla operativa (CODEX)")
            st.caption("RUBRO | ITEM | STOCK (día_0) | LUN | MAR | MIE | JUE | VIE | SAB | DOM")
            st.dataframe(sanitize_for_display(op_table))

    # Pestaña opcional de Notas
    _tabs = st.tabs(["Pedidos", "Reporte", "Notas"])
    with _tabs[2]:
        st.markdown(load_notes_md(), unsafe_allow_html=False)

    s_yesterday, found_y = yesterday_state(ventas_df, target_ts)
    if not found_y:
        st.info(f"No hay estado para {(target_ts - pd.Timedelta(days=1)).strftime('%d/%m/%Y')}; se usara '-- OK'.")

    if btn_run:
        with st.spinner("Calculando prediccion y pedido neto"):
            # Intento: usar estado de PRED_VENTA para la fecha
            pred_state = None
            try:
                xls_tmp = pd.ExcelFile(io.BytesIO(file_bytes))
                if "PRED_VENTA" in xls_tmp.sheet_names:
                    dfpv = xls_tmp.parse("PRED_VENTA", header=None)
                    fechas = parse_fecha_dmY(dfpv.iloc[2:13, 1])
                    estados = dfpv.iloc[2:13, 8].astype(str)
                    for d, s in zip(fechas, estados):
                        if pd.notna(d) and pd.Timestamp(d).normalize() == target_ts.normalize():
                            pred_state = s.strip()
                            break
            except Exception:
                pred_state = None
            if pred_state and pred_state in STATES:
                one_hot = pd.Series([1.0 if stn == pred_state else 0.0 for stn in STATES], index=STATES)
                # Reutilizar builder de pedidos para la fecha objetivo, respetando target/múltiplo
                pedido = build_weekly_orders(
                    Q,
                    file_bytes,
                    stock_df,
                    order_multiple=order_multiple,
                    target=target_stock,
                    enable_am=enable_am,
                    am_frac=am_frac,
                )
                # quedarnos solo con la col de la fecha
                col = target_ts.strftime("%d/%m/%Y")
                if not pedido.empty and col in pedido.columns:
                    pedido = pedido[["Rubro", "Item", col]].rename(columns={col: "Pedido sugerido (u)"})
            else:
                pedido = predict_for_date(ventas_df, diario_df, maps, probs_used, Q, stock_df, target_ts, s_yesterday, order_multiple=order_multiple)
        st.subheader("Pedido sugerido (fecha objetivo)")
        st.dataframe(sanitize_for_display(pedido))

        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            try:
                if "PRED_VENTA" in xls_view.sheet_names:
                    pv = xls_view.parse("PRED_VENTA", header=None)
                    pv_view = pv.iloc[1:13, 0:9]
                    pv_view.to_excel(writer, index=False, sheet_name="PREDICCION_DE_VENTA_SEMANAL")
            except Exception:
                pass
            try:
                if not weekly_grid.empty:
                    weekly_grid.to_excel(writer, index=False, sheet_name="PREDICCION_SEMANAL_ITEMS")
            except Exception:
                pass
            try:
                if not weekly_orders.empty:
                    weekly_orders.to_excel(writer, index=False, sheet_name="PEDIDO_SEMANAL")
            except Exception:
                pass
            try:
                # Agregar TABLA_OPERATIVA
                if 'op_table' in locals() and op_table is not None and not op_table.empty:
                    # Asegurar enteros >= 0
                    op_exp = op_table.copy()
                    for c in [col for col in op_exp.columns if c not in ("RUBRO", "ITEM")]:
                        op_exp[c] = pd.to_numeric(op_exp[c], errors="coerce").fillna(0).clip(lower=0).astype(int)
                    op_exp.to_excel(writer, index=False, sheet_name="TABLA_OPERATIVA")
            except Exception:
                pass
        file_name = f"MADAME_CEPHALPOD_ORACLE_{target_ts.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="Descargar Excel",
            data=out.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
except Exception as e:
    st.error(f"Error procesando el archivo: {e}")
    st.stop()
